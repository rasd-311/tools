#到期自動發email提醒 v0.1
import datetime
import win32com.client as win32
import openpyxl

#參數
wb = openpyxl.load_workbook("sample.xlsx")     # 開啟 Excel 檔案
names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
sheet = wb["data"]        # 取得工作表名稱為「工作表1」的內容
Email = []
Name = []
Staffno = []
LastDate = []

def send_mail(Name, Staffno, LastDate, mail_to, sign):    
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.To = mail_to  #收件人
    #mail.CC = "" #CC
    mail.Subject = "測試_機場証到期提醒"  #主旨
    date = LastDate.strftime("%Y-%m-%d")
    mail.Body = "Dear " + Name + "(" + str(Staffno) + ") ,\n\n" + "你的機場証在 " + date +"過期\n\nBest Regards\n"+sign
    mail.Send()       #發送

def read_xlsx(wb, mail_to, Name, Staffno, LastDate) :
    Today = datetime.datetime.today()
    print("now time : "+str(Today))
    sign = input("寄件人落款(如 : Daniel Ieong) : ") #寄件人
    for i in range(sheet.max_row-1) : #max_row 最大列數
        for j in range(sheet.max_column) : #max_column 最大行數
            v = sheet.cell(row=i+2, column=j+1)
            if j == 0 : #mail_to
                mail_to.append(v.value)
            if j == 1 : #Name         
                Name.append(v.value)
            if j == 2 : #Staffno
                Staffno.append(v.value)
            if j == 3 : #LastDate
                LastDate.append(v.value)
                days = (v.value-Today).days + 1
                if days <= 3 :
                    send_mail(Name[i], Staffno[i], LastDate[i], mail_to[i], sign)
    
read_xlsx(wb, Email, Name, Staffno, LastDate)