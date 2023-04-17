#到期自動發email提醒 v0.1
import datetime
import win32com.client as win32
import openpyxl
import os

#參數
wb = openpyxl.load_workbook("sample.xlsx")     # 開啟 Excel 檔案
names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
sheet = wb["cc"]        # 取得工作表名稱為「data」的內容
Email = []
Name = []
Staffno = []
LastDate = []
cc = ""

def read_cc(wb, cc):
    for i in range(sheet.max_row-1) : #max_row 最大列數
        for j in range(sheet.max_column) : #max_column 最大行數
            v = sheet.cell(row=i+2, column=j+1)
            if j == 0 :
                cc = cc + str(v.value) + ";"
    print(cc)
    return cc
cc = read_cc(wb, cc)

sheet = wb["data"]
def send_mail(Name, Staffno, LastDate, mail_to, days, cc):  
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.To = mail_to  #收件人
    mail.CC = cc
    mail.Subject = "測試_機場証到期提醒"  #主旨
    date = LastDate.strftime("%Y-%m-%d")
    mail.Body = "Dear " + Name + "(" + str(Staffno) + ") ,\n\n" + "請注意, 你的機場証在 " + date + " 過期\n若已更換, 請通知相關負責人更新記錄, 謝謝!"
    mail.Send()       #發送
    print("已發送給  "+Name+"("+str(Staffno)+")  提醒郵件")

def read_xlsx(wb, mail_to, Name, Staffno, LastDate, cc) :
    Today = datetime.datetime.today()
    print("NOW : "+str(Today))
    #sign = input("寄件人落款(如 : Daniel Ieong) : ") #寄件人
    for i in range(sheet.max_row-3) : #max_row 最大列數
        for j in range(sheet.max_column) : #max_column 最大行數
            v = sheet.cell(row=i+2, column=j+1)
            if j == 0 : #mail_to
                mail_to.append(v.value)
            if j == 1 : #Name         
                Name.append(v.value)
            if j == 2 : #Staffno
                Staffno.append(str(v.value))
            if j == 3 : #LastDate
                LastDate.append(v.value)
                days = (v.value-Today).days + 1
                if days <= 7 and days >= 6 :
                    send_mail(Name[i], Staffno[i], LastDate[i], mail_to[i], days, cc)
                    os.system("pause")
                if days <= 1 and days >= 0 :
                    send_mail(Name[i], Staffno[i], LastDate[i], mail_to[i], days, cc)
                    os.system("pause")
            
read_xlsx(wb, Email, Name, Staffno, LastDate, cc)
os.system("pause")