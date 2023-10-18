#到期自動發email提醒 v2.0
import datetime
import win32com.client as win32
import openpyxl
import os

#參數
wb = openpyxl.load_workbook("sample.xlsx")     # 開啟 Excel 檔案, Open Excel file
names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱, Read all worksheet names in Excel
Email = []
Name = []
Staffno = []
LastDate = []
cc = ""

def read_cc(wb):
    sheet = wb["cc"]        # 取得工作表名稱為「cc」的內容, Get the contents of the worksheet named "cc"
    cc = ""
    for i in range(sheet.max_row-1) : #max_row 最大列數
        for j in range(sheet.max_column) : #max_column 最大行數
            v = sheet.cell(row=i+2, column=j+1)
            if v.value is None :
                break
            if j == 0 :
                cc = cc + str(v.value) + ";"
    print("cc : " + cc)
    return cc
cc = read_cc(wb)

def send_mail(Name, Staffno, LastDate, mail_to, days, cc):  
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.SentOnBehalfOfName = "sender@test.com"  #寄件人 sender
    mail.To = mail_to  #收件人 receiver
    mail.CC = cc
    mail.Subject = "Airport access card expiry reminder_機場証到期提醒"  #主旨 Subject
    date = LastDate.strftime("%Y-%m-%d")
    mail.Body = "Dear " + Name + "(" + str(Staffno) + ") ,\n\n" + "請注意, 你的機場証在 " + date + " 過期\n若已更換, 請通知相關負責人更新記錄, 謝謝!"+"\n\nYour airport access card expired in "+ date + "\nif it has been replaced, please notify the relevant person in charge to update the record, thank you!"
    mail.Send()       #發送 sent

def read_xlsx(wb, mail_to, Name, Staffno, LastDate, cc) :
    sheet = wb["data"]
    Today = datetime.datetime.today()
    print("NOW : "+str(Today))
    r = sheet.max_row-1
    c = sheet.max_column
    for i in range(r) : #max_row 最大列數
        for j in range(c) : #max_column 最大行數
            v = sheet.cell(row=i+2, column=j+1)
            if v.value is None :
                break
            if j == 0 : #mail_to
                mail_to.append(v.value)
            if j == 1 : #Name         
                Name.append(v.value)
            if j == 2 : #Staffno
                Staffno.append(str(v.value))
            if j == 3 : #LastDate
                LastDate.append(v.value)
                days = (v.value-Today).days + 1
                if days <= 7 :
                    send_mail(Name[i], Staffno[i], LastDate[i], mail_to[i], days, cc)

read_xlsx(wb, Email, Name, Staffno, LastDate, cc)

def keepalive(mail_to, cc) :
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.SentOnBehalfOfName = "sender@test.com"  #寄件人 sender
    mail.To = mail_to  #收件人 receiver
    mail.Subject = "still alive"  #主旨 Subject
    current_dateTime = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    mail.Body = current_dateTime + "\nstill alive"
    mail.Send()       #發送
keepalive("receiver@test.com")