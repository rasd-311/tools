#到期自動發email提醒 v2.0
import win32com.client as win32
import openpyxl
from datetime import datetime

#參數
wb = openpyxl.load_workbook("sample.xlsx")     # 開啟 Excel 檔案
names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
Email = []
Name = []
Staffno = []
LastDate = []
cc = ""
keepalive = ""

def get_sheet(wb, sheet):
    sheet = wb[sheet]        # 取得工作表名稱為「sheet」的內容
    sheet_list = ""
    cal_null = 0 #判斷空行參數, 防止中間有空行報錯
    for sheet_row in range(sheet.max_row-1) : #max_row 最大列數
        for sheet_column in range(sheet.max_column) : #max_column 最大行數
            v = sheet.cell(row=sheet_row+2, column=sheet_column+1)
            if v.value is None :
                cal_null = cal_null + 1
                if cal_null == 99999: #判斷空行用, 防止中間有空行報錯
                    break
            if sheet_column == 0 :
                sheet_list = sheet_list + str(v.value) + ";"
    print("sheet_list : " + sheet_list)
    return sheet_list
cc = get_sheet(wb, "cc")
keepalive = get_sheet(wb, "keepalive")

def send_mail(Name, Staffno, LastDate, mail_to, days, cc):  
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.SentOnBehalfOfName = "License_Reminder@xxxxx.com"  #寄件人
    mail.To = mail_to  #收件人
    mail.CC = cc
    mail.Subject = "維修執照到期提醒"  #主旨
    date = LastDate.strftime("%Y-%m-%d")
    mail.Body = "Dear " + Name +",\n (" + str(Staffno) + ") ,\n\n" + "請注意, 你的執照在 " + date + " 過期\n若已更換, 請通知相關負責人更新記錄, 謝謝!"
    mail.Send()       #發送
    print("已發送給  "+Name+"("+str(Staffno)+")  提醒郵件")

def main(wb, mail_to, Name, Staffno, LastDate, cc) :
    skip_conut = 0 #計算跳過無效email
    sheet = wb["data"]
    Today = datetime.datetime.today()
    print("NOW : "+str(Today))
    for sheet_row in range(sheet.max_row) : #max_row 最大列數
        for sheet_column in range(4) : #max_column 最大行數
            v = sheet.cell(row=sheet_row + 1, column=sheet_column + 1)
            if sheet_column == 0 and ("@test.com" in v.value) == True : #mail_to, 判定是否為有效電郵地址
                mail_to.append(v.value)
            if sheet_column == 0 and ("@test.com" in v.value) == False: #如不是有效電郵地址, 跳過
                skip_conut = skip_conut + 1
                break
            if sheet_column == 1 : #Name     
                Name.append(v.value)
            if sheet_column == 2 : #Staffno
                Staffno.append(str(v.value))
            if sheet_column == 3 : #LastDate
                LastDate.append(v.value)
                days = (v.value-Today).days + 1
                if days <= 7 :
                    i = sheet_row - skip_conut
                    send_mail(Name[i], Staffno[i], LastDate[i], mail_to[i], days, cc)
                

main(wb, Email, Name, Staffno, LastDate, cc)

def keepalive_mail(keepalive) :
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.SentOnBehalfOfName = "keepalive@xxxxx.com"  #寄件人
    mail.To = keepalive  #收件人
    mail.Subject = "到期郵件提醒程序已執行"  #主旨
    current_dateTime = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    mail.Body = current_dateTime + "\nLicense_Reminder主機仍正常運行中\n到期郵件提醒程序已執行"
    mail.Send()       #發送

keepalive_mail(keepalive)
