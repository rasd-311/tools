#到期自動發email提醒 v0.1
import openpyxl
import os

#參數
wb = openpyxl.load_workbook("sample.xlsx")     # 開啟 Excel 檔案
names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
sheet = wb["data"]        # 取得工作表名稱為「data」的內容
#sheet_cc = wb["cc"]
Email = []
Name = []
Staffno = []
LastDate = []
cc = ""

def read_cc(wb, cc):
    for i in range(sheet.max_row-3) : #max_row 最大列數
        for j in range(sheet.max_column) : #max_column 最大行數
            v = sheet.cell(row=i+2, column=j+1)
            if j == 0 :
                print("i+2:"+str(i+2)+"j+1 : "+str(j+1)+", "+v.value)
                cc = cc + str(v.value) + ";"
            if j == 2 :
                print("i+2:"+str(i+2)+"j+1 : "+str(j+1)+", "+str(v.value))
                cc = cc + str(v.value) + ";"
            if j == 3 :
                print("i+2:"+str(i+2)+"j+1 : "+str(j+1)+", "+str(v.value))
                cc = cc + str(v.value) + ";"
    #print(cc)
    return cc
cc = read_cc(wb, cc)