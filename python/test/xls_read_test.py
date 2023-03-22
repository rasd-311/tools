import openpyxl
import datetime
wb = openpyxl.load_workbook("sample.xlsx")     # 開啟 Excel 檔案

names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
sheet = wb['工作表1']        # 取得工作表名稱為「工作表1」的內容
mail_to = []
Name = []
Staffno = []
LastDate = []
#max_row 最大列數、max_column 最大行數

def read_xlsx(wb, mail_to, Name, Staffno, LastDate) :
    Today = datetime.date.today()
    #print("today type : "+str(type(Today)))
    for i in range(sheet.max_column) :
        for j in range(sheet.max_row) :
            v = sheet.cell(row=i+1, column=j+1)
            #print("i = "+str(i)+", j = "+str(j)+", "+str(v.value))
            if i == 0 : #mail_to
                mail_to.append(v.value)
            if i == 1 : #name         
                Name.append(v.value)
            if i == 2 : #staffno
                Staffno.append(v.value)
            if i == 3 : #lastdate
                LastDate.append(v.value)
                #days = (Today-v.value).days
                #print("last date type : "+str(type(v.value)))
                
    print(mail_to)
    print(Name)
    print(Staffno)
    print(LastDate)

read_xlsx(wb, mail_to, Name, Staffno, LastDate)
'''
lastdate = "2018/08/09"
format = "%Y/%m/%d" #specifify the format of the date_string.
lastdate = datetime.datetime.strptime(date_string, format)
print(date)
'''
        