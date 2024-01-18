#到期自動發email提醒 v1.0
import win32com.client as win32
import openpyxl
from datetime import datetime

#參數
wb = openpyxl.load_workbook("sample.xlsx")     # 開啟 Excel 檔案
names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
Email = []
Name = []
Staffno = []
LastDate = []
cc = ""
keepalive = ""

def get_sheet(wb, sheet):
    sheet = wb[sheet]        # 取得工作表名稱為「sheet」的內容
    sheet_list = ""
    for i in range(sheet.max_row-1) : #max_row 最大列數
        for j in range(sheet.max_column) : #max_column 最大行數
            v = sheet.cell(row=i+2, column=j+1)
            if v.value is None :
                break
            if j == 0 :
                sheet_list = sheet_list + str(v.value) + ";"
    print("sheet_list : " + sheet_list)
    return sheet_list
cc = get_sheet(wb, "cc")
keepalive = get_sheet(wb, "keepalive")

def send_mail(Name, Staffno, LastDate, mail_to, days, cc):  
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.SentOnBehalfOfName = "License_Reminder@xxxxx.com"  #寄件人
    mail.To = mail_to  #收件人
    mail.CC = cc
    mail.Subject = "Aircraft Maintenance Personnel License expired reminder_維修執照到期提醒"  #主旨
    date = LastDate.strftime("%Y-%m-%d")
    mail.Body = "Dear " + Name +",\n (" + str(Staffno) + ") ,\n\n" + "請注意, 你的執照在 " + date + " 過期\n若已更換, 請通知相關負責人更新記錄, 謝謝!"+"\n\nYour A320 model endorsement for maintenance license expired in "+ date + "\nif it has been replaced, please notify the relevant person in charge to update the record, thank you!"
    mail.Send()       #發送
    print("已發送給  "+Name+"("+str(Staffno)+")  提醒郵件")

def main(wb, mail_to, Name, Staffno, LastDate, cc) :
    sheet = wb["data"]
    Today = datetime.today()
    print("NOW : "+str(Today))
    r = sheet.max_row-1
    c = sheet.max_column
    for i in range(r) : #max_row 最大列數
        for j in range(c) : #max_column 最大行數
            v = sheet.cell(row=i+2, column=j+1)
            if v.value is None :
                break
            if j == 0 : #mail_to
                mail_to.append(v.value)
            if j == 1 : #Name         
                Name.append(v.value)
            if j == 2 : #Staffno
                Staffno.append(str(v.value))
            if j == 3 : #LastDate
                LastDate.append(v.value)
                days = (v.value-Today).days + 1
                
                if days <= 7 and days >= 0 :
                    send_mail(Name[i], Staffno[i], LastDate[i], mail_to[i], days, cc)


main(wb, Email, Name, Staffno, LastDate, cc)

def keepalive_mail(keepalive) :
    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.SentOnBehalfOfName = "keepalive@xxxxx.com"  #寄件人
    mail.To = keepalive  #收件人
    mail.Subject = "已執行"  #主旨
    current_dateTime = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
    mail.Body = current_dateTime + "\nLicense_Reminder主機(192.168.18.40)仍正常運行中\n維修執照的A320机型签注到期郵件提醒程序已執行"
    mail.Send()       #發送

keepalive_mail(keepalive)